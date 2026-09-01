"""
TlachIA Metrics - openalex_indicators_engine
exporters/excel_builder.py
Generador de libros Excel estilizados y tipados con openpyxl.
"""
import os
import pandas as pd
from pathlib import Path
from typing import Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def save_styled_excel(df: pd.DataFrame, file_path: Union[str, Path], sheet_name: str = 'Metrics'):
    """Guarda un DataFrame como archivo Excel (.xlsx) con formato profesional."""
    p = Path(file_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    
    if df is None or len(df) == 0:
        df = pd.DataFrame({'Info': ['No data available for this criteria']})

    with pd.ExcelWriter(p, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Estilizado con openpyxl
    wb = openpyxl.load_workbook(p)
    ws = wb[sheet_name]

    # Paleta de colores TlachIA Metrics (Azul oscuro institucional y texto blanco)
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    ws.row_dimensions[1].height = 28

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        # Header style
        header_cell = col[0]
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment

        # Ajuste de ancho de columna
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Body cell borders & alignment
        for cell in col[1:]:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    wb.save(p)
